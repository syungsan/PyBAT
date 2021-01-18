#!/usr/bin/env python
# coding: utf-8

import openpyxl


def get_value_list(t_2d):
    return([[cell.value for cell in row] for row in t_2d])

def get_list_2d(filePath, sheetName, start_row, end_row, start_col, end_col):

    wb = openpyxl.load_workbook(filePath)
    sheet = wb[sheetName]

    return get_value_list(sheet.iter_rows(min_row=start_row,
                                          max_row=end_row,
                                          min_col=start_col,
                                          max_col=end_col))

def over_write_list_2d(filePath, sheetName, l_2d, start_row, start_col):

    wb = openpyxl.load_workbook(filePath)
    sheet = wb[sheetName]

    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                       column=start_col + x,
                       value=l_2d[y][x])

    wb.save(filePath)

def over_write_one_value(filePath, sheetName, value, cell):

    wb = openpyxl.load_workbook(filePath)
    sheet = wb[sheetName]
    sheet[cell] = value

    wb.save(filePath)

def over_write_list_1d(filePath, sheetName, l_1d, start_row, start_col):

    wb = openpyxl.load_workbook(filePath)
    sheet = wb[sheetName]

    for y, row in enumerate(l_1d):
        sheet.cell(row=start_row + y,
                   column=start_col,
                   value=row)

    wb.save(filePath)


if __name__ == '__main__':

    import pprint
    import shutil

    shutil.copy("../data/result_template.xlsx", "../temp/result_test.xlsx")

    test1_l_2d = get_list_2d("../temp/result_test.xlsx", "VAD", 3, 34, 1, 4)
    test2_l_2d = get_list_2d("../temp/result_test.xlsx", "VAD", 3, 34, 6, 9)

    pprint.pprint(test1_l_2d, width=40)
    pprint.pprint(test2_l_2d, width=40)

    test1_datas = [["か", 0.1, 0.2, 0.1], ["は", 0.3, 0.4, 0.1]]
    test2_datas = [["か", 0.5, 0.6, 0.1], ["は", 0.7, 0.8, 0.1]]

    over_write_list_2d("../temp/result_test.xlsx", "VAD", test1_datas, 3, 1)
    over_write_list_2d("../temp/result_test.xlsx", "VAD", test2_datas, 3, 6)

    wb = openpyxl.load_workbook("../temp/result_test.xlsx")
    sheet = wb["VAD"]

    pprint.pprint(list(sheet.values), width=40)
